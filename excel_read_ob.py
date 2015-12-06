__author__ = 'jessem'

import openpyxl as xl
from settings import settings
from openpyxl import utils
from m3_rest import M3_rest

WORKBOOK = '/home/jessem/PycharmProjects/excel_interface/MMS200MI.xlsx'
MAX_ROWS = settings['max_rows'] + 3
MAX_COLUMNS = settings['max_columns']
USER = settings['user']
PW = settings['password']

class excel_mi:

    def __init__(self,
                 file,
                 username,
                 password
                 ):
        self.file = file
        self.username = username
        self.password = password
        self.workbook = xl.load_workbook(self.file)
        self.base_url = settings['base_exec_url']
        self.max_columns = MAX_COLUMNS
        self.rows = MAX_ROWS
        self.transactions = self.get_excel_data()

    def get_excel_data(self):

        all_transactions = []

        for ws in self.workbook:
            program = ws.cell(row = 1, column = 1).value
            transaction = ws.cell(row = 2, column = 1).value

            transaction_grid = []
            last_field = None


            for row in ws.get_squared_range(2, 4, self.max_columns, self.rows):
                if row[0].value:
                    grid_index = row[0].row - 4
                    transaction_grid.append([])
                    for field in row:
                        if field.value and not last_field:
                            transaction_grid[grid_index].append(field.value)
                        elif not last_field:
                            last_field = utils.column_index_from_string(field.column) - 1
                        elif utils.column_index_from_string(field.column) <= last_field:
                            transaction_grid[grid_index].append(field.value)

            process_rows = []

            for i in range(len(transaction_grid[0])):
                for j in range(len(transaction_grid)):

                    if i == 0 and j > 0:
                        process_rows.append({})
                        process_rows[j-1]['excel_cell'] = {'column': 1, 'row': j + 4}
                    elif j == 0:
                        field_name = transaction_grid[j][i]
                    elif transaction_grid[j][i]:
                        process_rows[j - 1][field_name] = transaction_grid[j][i]


            for process_row in process_rows:
                cell_ref = process_row.pop('excel_cell', {})
                all_transactions.append({'program': program,
                                         'transaction': transaction,
                                         'parameters': process_row,
                                         'result_cell': cell_ref,
                                         'ws': ws
                                         })

        return all_transactions

    def process_trans(self):

        for mi in self.transactions:
            mi_object = M3_rest(self.base_url,
                                mi['program'],
                                mi['transaction'],
                                mi['parameters'],
                                self.username,
                                self.password)
            mi['result'] = mi_object.m3_get()
            ws = mi['ws']
            result_cell = ws.cell(row = mi['result_cell']['row'], column = mi['result_cell']['column'])
            result_cell.value = mi['result']

        self.workbook.save(filename=self.file)

    def __iter__(self):
        return iter(self.transactions)


def main():
    ex_trans = excel_mi(WORKBOOK, USER, PW)
    ex_trans.process_trans()

    for i in ex_trans.transactions:
        print(i)


if __name__ == '__main__':
    main()

