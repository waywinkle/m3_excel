__author__ = 'jessem'

import openpyxl as xl
from settings import settings
from openpyxl import utils

WORKBOOK = '/home/jessem/PycharmProjects/excel_interface/MMS200MI.xlsx'
MAX_ROWS = settings['max_rows'] + 3
MAX_COLUMNS = settings['max_columns']


def get_excel_data(workbook):
    wb = xl.load_workbook(workbook)
    all_transactions = []

    for ws in wb:
        program = ws.cell(row = 1, column = 1).value
        transaction = ws.cell(row = 2, column = 1).value

        transaction_grid = []
        last_field = None

        for row in ws.get_squared_range(2, 4, MAX_COLUMNS, MAX_ROWS):
            if row[0].value:
                grid_index = row[0].row - 4
                transaction_grid.append([])
                for field in row:
                    if field.value and not last_field:
                        transaction_grid[grid_index].append(field.value)
                    elif not last_field:
                        last_field = utils.column_index_from_string(field.column) - 1
                    elif utils.column_index_from_string(field.column) <= last_field:
                        transaction_grid[grid_index].append(field.value)

        process_rows = []

        for i in range(len(transaction_grid[0])):
            for j in range(len(transaction_grid)):

                if i == 0 and j > 0:
                    process_rows.append({})
                    process_rows[j-1]['excel_cell'] = {'column': 1, 'row': j + 3}
                elif j == 0:
                    field_name = transaction_grid[j][i]
                elif transaction_grid[j][i]:
                    process_rows[j - 1][field_name] = transaction_grid[j][i]


        for process_row in process_rows:
            cell_ref = {'ws':ws}
            cell_ref.update(process_row.pop('excel_cell', {}))
            all_transactions.append({'program': program,
                                     'transaction': transaction,
                                     'parameters': process_row,
                                     'result_cell': cell_ref
                                     })

    return all_transactions


def main():
    result = get_excel_data(WORKBOOK)

    for i in result:
        print(len(i), i)


if __name__ == '__main__':
    main()

