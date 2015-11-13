__author__ = 'jessem'

from openpyxl import Workbook
import m3_rest_getdetail
from settings import settings

FILE_NAME = 'AddItmBasic.xlsx'
PATH = settings['file_output_path']
PROGRAM = 'MMS200MI'
TRANSACTION = 'CpyItmBasic'


def create_empty_transaction(program, transaction):
    wb = Workbook()

    ws = wb.active
    ws.title = program + '_' + transaction

    mi_detail = m3_rest_getdetail.get_mi_detail(program, transaction)

    column = 1
    ws.cell(row = 1, column = column, value=program)
    ws.cell(row = 2, column = column, value=transaction)

    for field in mi_detail:
        row = 3
        for value in field:
            active_cell = ws.cell(row = row, column = column)
            active_cell.value = value
            row += 1
        column += 1

    file = PATH + '/' + program + '_' + transaction + '.xlsx'
    wb.save(filename=file)


def main():
    create_empty_transaction(PROGRAM, TRANSACTION)


if __name__ == '__main__':
    main()